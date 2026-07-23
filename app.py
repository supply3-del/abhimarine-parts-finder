"""
AbhiMarine Parts Finder  -  team tool for answering customer enquiries fast.
Search 41k+ marine engine / equipment spares, see model-group interchangeability,
info-richness star ratings, build an enquiry -> quote shortlist, and export to
Excel / PDF / WhatsApp.

Run:   streamlit run app.py
Login: shared team password (see TEAM_PASSWORD below / .streamlit/secrets.toml)
"""

import io
import os
import sqlite3
import urllib.parse
from datetime import datetime

import pandas as pd
import streamlit as st

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "parts.db")
TEAL, TEAL_DK, INK, MUTE = "#0F9D8F", "#0B6B61", "#0B2E2A", "#5B756F"

# Shared team password. Override with .streamlit/secrets.toml -> TEAM_PASSWORD = "..."
TEAM_PASSWORD = st.secrets.get("TEAM_PASSWORD", "abhimarine") if hasattr(st, "secrets") else "abhimarine"

st.set_page_config(page_title="AbhiMarine Parts Finder", page_icon="⚓",
                   layout="wide", initial_sidebar_state="expanded")

# ----------------------------------------------------------------- styling
st.markdown(f"""
<style>
:root {{ --teal:{TEAL}; --teal-dk:{TEAL_DK}; --ink:{INK}; }}
.block-container {{ padding-top:1.4rem; padding-bottom:7rem; max-width:1400px; }}
#MainMenu, footer {{ visibility:hidden; }}
h1,h2,h3 {{ color:{INK}; letter-spacing:-.01em; }}
.hero {{ background:linear-gradient(120deg,{TEAL} 0%,{TEAL_DK} 100%);
        color:#fff; padding:1.2rem 1.5rem; border-radius:16px; margin-bottom:1rem; }}
.hero h1 {{ color:#fff; margin:0; font-size:1.7rem; }}
.hero p {{ color:#d6f3ee; margin:.2rem 0 0; font-size:.92rem; }}
.kpi {{ background:#F1FAF8; border:1px solid #d9efeb; border-radius:14px;
        padding:.8rem 1rem; text-align:center; }}
.kpi .v {{ font-size:1.5rem; font-weight:700; color:{TEAL_DK}; }}
.kpi .l {{ font-size:.72rem; color:{MUTE}; text-transform:uppercase; letter-spacing:.04em; }}
.card {{ background:#fff; border:1px solid #e6efed; border-left:5px solid {TEAL};
         border-radius:12px; padding:.85rem 1rem; margin-bottom:.6rem;
         box-shadow:0 1px 3px rgba(11,46,42,.05); }}
.card:hover {{ box-shadow:0 4px 14px rgba(15,157,143,.16); border-left-color:{TEAL_DK}; }}
.pn {{ font-weight:700; color:{INK}; font-size:1.02rem; }}
.meta {{ color:{MUTE}; font-size:.82rem; margin-top:.15rem; }}
.pill {{ display:inline-block; background:#E3F5F2; color:{TEAL_DK}; border-radius:20px;
         padding:.12rem .6rem; font-size:.72rem; font-weight:600; margin-right:.3rem; }}
.pill.warn {{ background:#FDECEC; color:#C0392B; }}
.pill.grey {{ background:#EEF2F1; color:{MUTE}; }}
.stars {{ color:#F2A900; font-size:.95rem; letter-spacing:1px; }}
.footbar {{ position:fixed; left:0; right:0; bottom:0; z-index:999;
           background:{INK}; color:#fff; padding:.7rem 1.5rem;
           box-shadow:0 -4px 18px rgba(0,0,0,.18); }}
.footbar b {{ color:#7BE0D3; }}
.loginwrap {{ max-width:420px; margin:6vh auto 0; }}
div[data-testid="stDataFrame"] {{ border-radius:10px; overflow:hidden; }}
.stButton>button {{ border-radius:10px; border:1px solid {TEAL};
        color:{TEAL_DK}; font-weight:600; }}
.stButton>button:hover {{ background:{TEAL}; color:#fff; }}
.stDownloadButton>button, .stLinkButton>a {{ background:{TEAL}; color:#fff !important;
        border:0; border-radius:10px; font-weight:600; }}
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------------------- login gate
def require_login():
    if st.session_state.get("authed"):
        return True
    st.markdown('<div class="loginwrap">', unsafe_allow_html=True)
    st.markdown('<div class="hero"><h1>⚓ AbhiMarine Parts Finder</h1>'
                '<p>Team access — please sign in.</p></div>', unsafe_allow_html=True)
    pwd = st.text_input("Team password", type="password")
    if st.button("Sign in", use_container_width=True):
        if pwd == TEAM_PASSWORD:
            st.session_state["authed"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.caption("Ask your admin for the shared team password.")
    st.markdown('</div>', unsafe_allow_html=True)
    return False


if not require_login():
    st.stop()

# ----------------------------------------------------------------- data
@st.cache_resource
def get_con():
    return sqlite3.connect(DB_PATH, check_same_thread=False)


@st.cache_data
def filter_options():
    con = get_con()
    def col(c):
        return [r[0] for r in con.execute(
            f"SELECT DISTINCT {c} FROM parts WHERE {c}!='' ORDER BY 1").fetchall()]
    return {"brands": col("brand"), "categories": col("category"),
            "conditions": col("condition"), "locations": col("location")}


@st.cache_data
def models_in_group(group_code):
    con = get_con()
    return [r[0] for r in con.execute(
        "SELECT DISTINCT model FROM model_groups WHERE group_code=? AND model!='' ORDER BY 1",
        (group_code,)).fetchall()]


@st.cache_data
def fuzzy_corpus():
    con = get_con()
    rows = con.execute(
        "SELECT row_id, part_name||' '||model||' '||brand||' '||part_number FROM parts"
    ).fetchall()
    return [r[0] for r in rows], [r[1] for r in rows]


def stars_html(n):
    return f"<span class='stars'>{'★'*int(n)}{'☆'*(5-int(n))}</span>"


def fuzzy_ids(text, limit=400):
    try:
        from rapidfuzz import process, fuzz
        ids, choices = fuzzy_corpus()
        hits = process.extract(text, choices, scorer=fuzz.WRatio,
                               limit=limit, score_cutoff=62)
        return [ids[i] for _, _, i in hits]
    except Exception:
        con = get_con()
        like = f"%{text}%"
        return [r[0] for r in con.execute(
            "SELECT row_id FROM parts WHERE part_name LIKE ? OR model LIKE ? LIMIT ?",
            (like, like, limit)).fetchall()]


def run_search(text, brands, conds, cats, locs, min_stars, in_stock):
    con = get_con()
    where, params, fuzzy = [], [], False
    if text.strip():
        terms = " ".join(f'"{t}"*' for t in text.split())
        ids = [r[0] for r in con.execute(
            "SELECT rowid FROM parts_fts WHERE parts_fts MATCH ? LIMIT 4000",
            (terms,)).fetchall()]
        if not ids:                       # exact search empty -> fuzzy fallback
            ids, fuzzy = fuzzy_ids(text), True
        if not ids:
            return pd.DataFrame(), False
        where.append(f"row_id IN ({','.join('?'*len(ids))})")
        params += ids
    if brands:
        where.append(f"brand IN ({','.join('?'*len(brands))})"); params += brands
    if conds:
        where.append(f"condition IN ({','.join('?'*len(conds))})"); params += conds
    if cats:
        where.append(f"category IN ({','.join('?'*len(cats))})"); params += cats
    if locs:
        where.append(f"location IN ({','.join('?'*len(locs))})"); params += locs
    if min_stars > 1:
        where.append("stars >= ?"); params.append(min_stars)
    if in_stock:
        where.append("available = 1")
    sql = "SELECT * FROM parts"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY stars DESC, available DESC, part_name LIMIT 1500"
    return pd.read_sql_query(sql, con, params=params), fuzzy


def group_summary(group_code):
    con = get_con()
    return pd.read_sql_query(
        "SELECT * FROM parts WHERE group_code=? ORDER BY available DESC, stars DESC, part_name",
        con, params=(group_code,))


# ----------------------------------------------------------------- shortlist
def add_to_shortlist(r):
    sl = st.session_state.setdefault("shortlist", {})
    rid = int(r["row_id"])
    if rid not in sl:
        sl[rid] = {"part_name": r["part_name"], "brand": r["brand"], "model": r["model"],
                   "group_code": r["group_code"], "condition": r["condition"],
                   "qty_avail": int(r["qty"]), "unit": r["unit"], "location": r["location"],
                   "rack": r["rack"], "part_number": r["part_number"],
                   "stars": int(r["stars"]), "qty_needed": 1}


# ----------------------------------------------------------------- exports
EXPORT_COLS = {"part_name": "Part Name", "part_number": "Part No.", "brand": "Brand",
               "model": "Model", "group_code": "Group", "condition": "Condition",
               "qty": "Qty", "unit": "Unit", "location": "Location", "rack": "Rack",
               "details": "Details / Dimension", "stars": "Info ★", "available": "In Stock"}


def style_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="0F9D8F")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def to_excel(df, group_code):
    sub = df[list(EXPORT_COLS)].rename(columns=EXPORT_COLS).copy()
    sub["In Stock"] = sub["In Stock"].map({1: "Yes", 0: "No"})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        sub.to_excel(xl, index=False, sheet_name="Spares")
        ws = xl.sheets["Spares"]; style_header(ws)
        for i, w in enumerate([34, 16, 14, 14, 12, 14, 7, 8, 14, 12, 40, 8, 9], 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf.seek(0); return buf.getvalue()


def quote_to_excel(items):
    df = pd.DataFrame(items.values())
    cols = ["part_name", "part_number", "brand", "model", "group_code", "condition",
            "qty_needed", "qty_avail", "unit", "location", "rack", "stars"]
    head = ["Part Name", "Part No.", "Brand", "Model", "Group", "Condition",
            "Qty Needed", "Qty Avail", "Unit", "Location", "Rack", "Info ★"]
    df = df[cols]; df.columns = head
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        df.to_excel(xl, index=False, sheet_name="Enquiry")
        ws = xl.sheets["Enquiry"]; style_header(ws)
        for i, w in enumerate([34, 16, 14, 14, 12, 14, 11, 10, 8, 14, 12, 8], 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf.seek(0); return buf.getvalue()


def _pdf(rows, head, colw, title, subtitle):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12*mm,
                            rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    ss = getSampleStyleSheet()
    ts = ParagraphStyle("t", parent=ss["Title"], textColor=colors.HexColor(TEAL_DK), fontSize=18)
    sb = ParagraphStyle("s", parent=ss["Normal"], textColor=colors.HexColor(MUTE), fontSize=9)
    cell = ParagraphStyle("c", parent=ss["Normal"], fontSize=7, leading=8)
    flow = [Paragraph(title, ts)]
    for line in subtitle:
        flow.append(Paragraph(line, sb))
    flow.append(Spacer(1, 6*mm))
    data = [head] + [[Paragraph(str(c), cell) if not isinstance(c, int) else c for c in row]
                     for row in rows]
    t = Table(data, repeatRows=1, colWidths=[w*mm for w in colw])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(TEAL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1FAF8")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cfe4e0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 7)]))
    flow.append(t); doc.build(flow); buf.seek(0); return buf.getvalue()


def to_pdf(df, group_code, models):
    rows = [[r["part_name"], r["part_number"], r["brand"], r["model"], r["condition"],
             int(r["qty"]), r["location"], r["rack"], int(r["stars"])]
            for _, r in df.iterrows()]
    return _pdf(rows, ["Part Name", "Part No.", "Brand", "Model", "Condition",
                       "Qty", "Location", "Rack", "★"],
                [70, 26, 22, 24, 24, 12, 26, 30, 10],
                f"AbhiMarine — Spares for Model Group {group_code}",
                ["Interchangeable models: " + ", ".join(models[:25]) + (" …" if len(models) > 25 else ""),
                 f"{len(df)} line items · generated {datetime.now():%d %b %Y %H:%M}"])


def quote_to_pdf(items):
    rows = [[v["part_name"], v["part_number"], v["brand"], v["model"], v["condition"],
             v["qty_needed"], v["qty_avail"], v["location"], v["rack"]]
            for v in items.values()]
    return _pdf(rows, ["Part Name", "Part No.", "Brand", "Model", "Condition",
                       "Qty Needed", "Qty Avail", "Location", "Rack"],
                [62, 24, 20, 22, 22, 18, 16, 26, 30],
                "AbhiMarine — Customer Enquiry / Quote",
                [f"{len(items)} line items · generated {datetime.now():%d %b %Y %H:%M}",
                 "AbhiMarine Pvt. Ltd. · marketing.team@abhimarine.com · abhimarine.com"])


def share_text(items):
    lines = ["*AbhiMarine — Parts enquiry*", ""]
    for v in items.values():
        pn = f" (PN {v['part_number']})" if v["part_number"] else ""
        lines.append(f"• {v['part_name']}{pn} — {v['brand']} {v['model']} · "
                     f"need {v['qty_needed']}, {v['qty_avail']} in stock · {v['condition']}")
    lines += ["", "AbhiMarine Pvt. Ltd. · abhimarine.com"]
    return "\n".join(lines)


# ----------------------------------------------------------------- header
con = get_con()
total = con.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
instock = con.execute("SELECT COUNT(*) FROM parts WHERE available=1").fetchone()[0]
ngroups = con.execute("SELECT COUNT(DISTINCT group_code) FROM parts WHERE group_code!=''").fetchone()[0]
nbrands = con.execute("SELECT COUNT(DISTINCT brand) FROM parts").fetchone()[0]

top = st.columns([6, 1])
top[0].markdown("""<div class="hero"><h1>⚓ AbhiMarine Parts Finder</h1>
<p>Search the spares inventory by model, brand, part name or part number — with
model-group interchangeability built in.</p></div>""", unsafe_allow_html=True)
if top[1].button("Sign out"):
    st.session_state.clear(); st.rerun()

k = st.columns(4)
for col, v, l in zip(k, [f"{total:,}", f"{instock:,}", f"{ngroups:,}", f"{nbrands:,}"],
                     ["Total Parts", "In Stock", "Model Groups", "Brands"]):
    col.markdown(f"<div class='kpi'><div class='v'>{v}</div><div class='l'>{l}</div></div>",
                 unsafe_allow_html=True)

st.session_state.setdefault("group", None)
st.session_state.setdefault("shortlist", {})

# ----------------------------------------------------------------- sidebar
opts = filter_options()
with st.sidebar:
    st.markdown("### 🔎 Filters")
    f_brand = st.multiselect("Brand", opts["brands"])
    f_cond = st.multiselect("Condition", opts["conditions"])
    f_cat = st.multiselect("Category", opts["categories"])
    f_loc = st.multiselect("Location", opts["locations"])
    f_stars = st.slider("Minimum info ★", 1, 5, 1)
    f_stock = st.checkbox("In-stock only", value=True)
    st.caption("Stars rate how complete a listing is — part number, dimensions, "
               "markings, genuinity and photo all add detail a buyer can trust.")

n_sl = len(st.session_state["shortlist"])
tab_search, tab_quote = st.tabs(["🔎 Search", f"🧾 Enquiry / Quote ({n_sl})"])

# ================================================================= SEARCH TAB
with tab_search:
    q = st.text_input("Search", placeholder="e.g.  L20 plunger   ·   MAN B&W scraper ring   ·   8525-784",
                      label_visibility="collapsed")
    res, fuzzy = run_search(q, f_brand, f_cond, f_cat, f_loc, f_stars, f_stock)

    msg = f"**{len(res):,}** matching parts" + (" (showing first 1,500)" if len(res) == 1500 else "")
    if fuzzy:
        msg += "  ·  _no exact match — showing closest results_"
    st.markdown(msg)

    if res.empty:
        st.info("No parts match. Try fewer filters or a broader search term.")
    else:
        grp_counts = res[res["group_code"] != ""]["group_code"].value_counts().head(40)
        if len(grp_counts):
            choices = ["— select a model group —"] + [f"{g}  ({n} here)" for g, n in grp_counts.items()]
            pick = st.selectbox("📦 View all interchangeable spares for a model group:",
                                choices, key="grp_pick")
            st.session_state["group"] = None if pick.startswith("—") else pick.split("  (")[0]

        for _, r in res.head(60).iterrows():
            c_card, c_btn = st.columns([12, 1])
            avail = ("<span class='pill'>In stock</span>" if r["available"]
                     else "<span class='pill warn'>Not available</span>")
            grp = f"<span class='pill grey'>{r['group_code']}</span>" if r["group_code"] else ""
            photo = "📷" if r["photo"] else ""
            c_card.markdown(f"""<div class="card">
            <div class="pn">{r['part_name']} {photo}</div>
            <div class="meta">{r['brand']} · {r['model'] or '—'} · Qty <b>{int(r['qty'])}</b> {r['unit']}
            · 📍 {r['location']} / {r['rack'] or '—'}{' · PN ' + r['part_number'] if r['part_number'] else ''}</div>
            <div style="margin-top:.4rem">{grp}<span class='pill grey'>{r['condition']}</span>{avail}
            {stars_html(r['stars'])}</div>
            {('<div class="meta" style="margin-top:.35rem">'+r['details'][:160]+'</div>') if r['details'] else ''}
            </div>""", unsafe_allow_html=True)
            c_btn.button("➕", key=f"add_{int(r['row_id'])}", help="Add to enquiry",
                         on_click=add_to_shortlist, args=(r,))

        if len(res) > 60:
            with st.expander(f"Show all {len(res):,} matches as a table"):
                st.dataframe(res[["part_name", "brand", "model", "group_code", "condition",
                                  "qty", "unit", "location", "rack", "part_number", "stars"]],
                             use_container_width=True, hide_index=True)

# ================================================================= QUOTE TAB
with tab_quote:
    sl = st.session_state["shortlist"]
    if not sl:
        st.info("Your enquiry list is empty. Add parts from the Search tab with the ➕ button.")
    else:
        st.markdown(f"### 🧾 Enquiry list — {len(sl)} part(s)")
        for rid, v in list(sl.items()):
            c1, c2, c3 = st.columns([8, 2, 1])
            c1.markdown(f"**{v['part_name']}** · {v['brand']} {v['model']} · "
                        f"{v['condition']} · {v['qty_avail']} in stock"
                        f"{' · PN ' + v['part_number'] if v['part_number'] else ''}")
            v["qty_needed"] = c2.number_input("Qty needed", min_value=1, value=v["qty_needed"],
                                              key=f"qn_{rid}", label_visibility="collapsed")
            if c3.button("🗑", key=f"rm_{rid}", help="Remove"):
                del sl[rid]; st.rerun()

        st.divider()
        e1, e2, e3, e4 = st.columns([1, 1, 1, 1])
        e1.download_button("⬇ Excel quote", quote_to_excel(sl),
                           file_name=f"AbhiMarine_enquiry_{datetime.now():%Y%m%d}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
        try:
            e2.download_button("⬇ PDF quote", quote_to_pdf(sl),
                               file_name=f"AbhiMarine_enquiry_{datetime.now():%Y%m%d}.pdf",
                               mime="application/pdf", use_container_width=True)
        except Exception:
            e2.caption("PDF: `pip install reportlab`")
        txt = share_text(sl)
        e3.link_button("💬 WhatsApp", "https://wa.me/?text=" + urllib.parse.quote(txt),
                       use_container_width=True)
        e4.link_button("✉ Email", "mailto:?subject=" + urllib.parse.quote("AbhiMarine parts enquiry") +
                       "&body=" + urllib.parse.quote(txt), use_container_width=True)
        if st.button("Clear list"):
            st.session_state["shortlist"] = {}; st.rerun()
        with st.expander("Preview share message"):
            st.code(txt)

# ----------------------------------------------------------------- bottom bar
g = st.session_state.get("group")
if g:
    gdf = group_summary(g)
    avail_df = gdf[gdf["available"] == 1]
    models = models_in_group(g)
    tot_qty = int(avail_df["qty"].sum())

    st.markdown(f"""<div class="footbar">
    Model group <b>{g}</b> — <b>{len(avail_df)}</b> spares in stock
    · total qty <b>{tot_qty}</b> · interchangeable across <b>{len(models)}</b> models</div>
    """, unsafe_allow_html=True)

    with st.expander(f"📦  {g} — full spares list & export", expanded=False):
        st.markdown("**Interchangeable models:** " +
                    " ".join(f"<span class='pill'>{m}</span>" for m in models),
                    unsafe_allow_html=True)
        c1, c2, c3 = st.columns([1, 1, 4])
        c1.download_button("⬇ Excel", to_excel(gdf, g),
                           file_name=f"AbhiMarine_{g.replace('/','-')}_spares.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
        try:
            c2.download_button("⬇ PDF", to_pdf(gdf, g, models),
                               file_name=f"AbhiMarine_{g.replace('/','-')}_spares.pdf",
                               mime="application/pdf", use_container_width=True)
        except Exception:
            c2.caption("PDF: `pip install reportlab`")
        c3.caption(f"{len(gdf)} total line items · {len(avail_df)} in stock")
        st.dataframe(
            gdf[["part_name", "brand", "model", "condition", "qty", "unit",
                 "location", "rack", "part_number", "stars", "available"]]
            .rename(columns={"available": "in_stock"}),
            use_container_width=True, hide_index=True)
