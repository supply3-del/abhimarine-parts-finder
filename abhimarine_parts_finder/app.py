"""
AbhiMarine Parts Finder  -  team tool for answering customer enquiries fast.
Search 41k+ marine engine / equipment spares, see model-group interchangeability,
info-richness star ratings, build an enquiry -> quote shortlist, and export to
Excel / PDF / WhatsApp.

Run:   streamlit run app.py
Login: shared team password (see TEAM_PASSWORD below / .streamlit/secrets.toml)
"""

import io
import os
import sqlite3
import urllib.parse
from datetime import datetime

import pandas as pd
import streamlit as st

from inquiry_parser import parse_inquiry, parse_pasted_text
from inquiry_match import (build_corpus, match_row, known_brands_models,
                           extract_brand_model, model_group_code)
from quote_templates import build_estimate_xlsx, build_estimate_pdf

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "parts.db")
TEAL, TEAL_DK, INK, MUTE = "#0F9D8F", "#0B6B61", "#0B2E2A", "#5B756F"

# Shared team password. Override with .streamlit/secrets.toml -> TEAM_PASSWORD = "..."
try:
    TEAM_PASSWORD = st.secrets.get("TEAM_PASSWORD", "abhimarine")
except Exception:
    TEAM_PASSWORD = "abhimarine"

st.set_page_config(page_title="AbhiMarine Parts Finder", page_icon="⚓",
                   layout="wide", initial_sidebar_state="expanded")

# ----------------------------------------------------------------- styling
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
:root {{ --teal:{TEAL}; --teal-dk:{TEAL_DK}; --ink:{INK}; --mute:{MUTE};
        --line:#E4EEEC; --surface:#FFFFFF; --bg:#F6FAF9;
        --shadow-sm:0 1px 2px rgba(11,46,42,.06); --shadow-md:0 4px 16px rgba(11,46,42,.08); }}

html, body, [class*="css"], .stMarkdown, input, button, textarea, select {{
    font-family:'Inter',-apple-system,'Segoe UI',Roboto,sans-serif; }}
.stApp {{ background:var(--bg); }}
* {{ transition:background-color .15s ease, border-color .15s ease, box-shadow .15s ease, color .15s ease; }}
@media (prefers-reduced-motion: reduce) {{ * {{ transition:none !important; }} }}

/* fuller-screen working area */
.block-container {{ padding-top:1.1rem; padding-bottom:7rem;
    padding-left:2.2rem; padding-right:2.2rem; max-width:1800px; }}
#MainMenu, footer, [data-testid="stToolbar"] {{ visibility:hidden; }}
h1,h2,h3 {{ color:{INK}; letter-spacing:-.015em; font-weight:700; }}
h3 {{ font-size:1.05rem; }}
/* tabular figures so numeric columns/KPIs don't jitter */
.kpi .v, [data-testid="stDataFrame"], [data-testid="stDataEditor"] {{ font-variant-numeric:tabular-nums; }}

.hero {{ background:linear-gradient(120deg,{TEAL} 0%,{TEAL_DK} 100%);
        color:#fff; padding:1.2rem 1.5rem; border-radius:16px; margin-bottom:1rem; }}
.hero h1 {{ color:#fff; margin:0; font-size:1.7rem; }}
.hero p {{ color:#d6f3ee; margin:.2rem 0 0; font-size:.92rem; }}

/* app header bar (replaces the hero banner); stHeader is empty here - toolbar
   is hidden above - so we drop it and stick our own bar to the top instead */
[data-testid="stHeader"] {{ display:none; }}
.st-key-appbar {{ position:sticky; top:0; z-index:999; background:var(--bg);
    border-bottom:1px solid var(--line); padding:.35rem 0 .3rem; margin-bottom:.9rem; }}
.brand {{ font-size:1.35rem; font-weight:700; color:{INK}; letter-spacing:-.02em; }}
.brand span {{ color:{TEAL_DK}; }}

.kpi {{ background:#F1FAF8; border:1px solid #d9efeb; border-radius:14px;
        padding:.8rem 1rem; text-align:center; }}
.kpi .v {{ font-size:1.5rem; font-weight:700; color:{TEAL_DK}; }}
.kpi .l {{ font-size:.72rem; color:{MUTE}; text-transform:uppercase; letter-spacing:.04em; }}

.card {{ background:var(--surface); border:1px solid var(--line); border-left:4px solid {TEAL};
         border-radius:10px; padding:.85rem 1rem; margin-bottom:.55rem; box-shadow:var(--shadow-sm); }}
.card:hover {{ box-shadow:var(--shadow-md); border-left-color:{TEAL_DK}; }}
.pn {{ font-weight:700; color:{INK}; font-size:1rem; }}
.meta {{ color:{MUTE}; font-size:.82rem; margin-top:.15rem; }}
.pill {{ display:inline-block; background:#E3F5F2; color:{TEAL_DK}; border-radius:6px;
         padding:.14rem .55rem; font-size:.7rem; font-weight:600; margin-right:.3rem; }}
.pill.warn {{ background:#FDECEC; color:#C0392B; }}
.pill.grey {{ background:#EEF2F1; color:{MUTE}; }}
.stars {{ color:#F2A900; font-size:.95rem; letter-spacing:1px; }}

.footbar {{ position:fixed; left:0; right:0; bottom:0; z-index:999;
           background:{INK}; color:#fff; padding:.7rem 2.2rem;
           box-shadow:0 -4px 18px rgba(0,0,0,.18); font-size:.9rem; }}
.footbar b {{ color:#7BE0D3; }}
.loginwrap {{ max-width:400px; margin:8vh auto 0; }}

.toolbar {{ background:#fff; border:1px solid #e6efed; border-radius:14px;
           padding:.7rem 1rem .1rem; margin-bottom:.9rem; box-shadow:0 1px 3px rgba(11,46,42,.05); }}

/* inputs & dropdowns: clear border on every text/search field and select */
[data-baseweb="input"], [data-baseweb="select"] > div,
.stTextInput input, .stNumberInput input, .stTextArea textarea,
[data-baseweb="select"] [data-baseweb="input"] {{
    border-radius:8px !important; border:1.5px solid #BFD4CF !important; background:#fff !important; }}
[data-baseweb="input"]:hover, [data-baseweb="select"] > div:hover,
.stTextInput input:hover, .stTextArea textarea:hover {{ border-color:{TEAL} !important; }}
[data-baseweb="input"]:focus-within, [data-baseweb="select"] > div:focus-within,
.stTextInput input:focus, .stTextArea textarea:focus {{
    border-color:{TEAL} !important; box-shadow:0 0 0 3px rgba(15,157,143,.15) !important; }}

/* tabs: pro software segmented look with teal active underline */
.stTabs [data-baseweb="tab-list"] {{ gap:.25rem; border-bottom:1px solid var(--line); }}
.stTabs [data-baseweb="tab"] {{ height:2.6rem; padding:0 1rem; font-weight:600; color:{MUTE};
        border-radius:8px 8px 0 0; }}
.stTabs [data-baseweb="tab"]:hover {{ color:{INK}; background:#EEF6F4; }}
.stTabs [aria-selected="true"] {{ color:{TEAL_DK} !important; }}
.stTabs [data-baseweb="tab-highlight"] {{ background:{TEAL}; height:3px; border-radius:3px; }}

/* data grids: framed, crisp header */
div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] {{
    border-radius:10px; overflow:hidden; border:1px solid var(--line); box-shadow:var(--shadow-sm); }}

/* buttons */
.stButton>button {{ border-radius:8px; border:1px solid {TEAL}; color:{TEAL_DK};
        font-weight:600; padding:.4rem 1rem; cursor:pointer; }}
.stButton>button:hover {{ background:{TEAL}; color:#fff; border-color:{TEAL};
        box-shadow:0 4px 12px rgba(15,157,143,.25); }}
.stButton>button:focus {{ box-shadow:0 0 0 3px rgba(15,157,143,.25); }}
.stDownloadButton>button, .stLinkButton>a {{ background:{TEAL}; color:#fff !important;
        border:0; border-radius:8px; font-weight:600; cursor:pointer; }}
.stDownloadButton>button:hover, .stLinkButton>a:hover {{ background:{TEAL_DK};
        box-shadow:0 4px 12px rgba(15,157,143,.3); }}

/* sidebar */
[data-testid="stSidebar"] {{ background:var(--surface); border-right:1px solid var(--line); }}

/* matching-inventory expanders: table flush to the box on all four sides, and
   pulled up tight under the header. Scoped to the right-hand match column
   (st.container key="matchcol") so the bottom spares-list expander — which also
   holds a dataframe but needs its padding for the buttons/pills — is untouched. */
.st-key-matchcol [data-testid="stExpander"] {{ margin-top:-.4rem !important; }}
.st-key-matchcol [data-testid="stExpanderDetails"] {{ padding:0 !important; }}
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------------------- login gate
def require_login():
    if st.session_state.get("authed"):
        return True
    st.markdown('<div class="loginwrap">', unsafe_allow_html=True)
    st.markdown('<div class="hero"><h1>⚓ AbhiMarine Parts Finder</h1>'
                '<p>Team access — please sign in.</p></div>', unsafe_allow_html=True)
    pwd = st.text_input("Team password", type="password")
    if st.button("Sign in", use_container_width=True):
        if pwd == TEAM_PASSWORD:
            st.session_state["authed"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.caption("Ask your admin for the shared team password.")
    st.markdown('</div>', unsafe_allow_html=True)
    return False


if not require_login():
    st.stop()

# ----------------------------------------------------------------- data
@st.cache_resource
def get_con():
    return sqlite3.connect(DB_PATH, check_same_thread=False)


@st.cache_data
def filter_options():
    con = get_con()
    def col(c):
        return [r[0] for r in con.execute(
            f"SELECT DISTINCT {c} FROM parts WHERE {c}!='' ORDER BY 1").fetchall()]
    groups = [r[0] for r in con.execute(
        "SELECT DISTINCT group_code FROM parts WHERE group_code!='' "
        "AND group_code!='UNKNOWN' ORDER BY 1").fetchall()]
    return {"brands": col("brand"), "categories": col("category"),
            "conditions": col("condition"), "locations": col("location"),
            "groups": groups}


@st.cache_data
def models_in_group(group_code):
    con = get_con()
    return [r[0] for r in con.execute(
        "SELECT DISTINCT model FROM model_groups WHERE group_code=? AND model!='' ORDER BY 1",
        (group_code,)).fetchall()]


@st.cache_data
def fuzzy_corpus():
    con = get_con()
    rows = con.execute(
        "SELECT row_id, part_name||' '||model||' '||brand||' '||part_number FROM parts"
    ).fetchall()
    return [r[0] for r in rows], [r[1] for r in rows]


@st.cache_data
def inquiry_corpus():
    return build_corpus(get_con())


@st.cache_data
def inquiry_terms():
    return known_brands_models(get_con())


def annotate_brand_model(df):
    """Normalize each row's brand/model to the inventory's own brand/model
    names. Uses the customer file's Maker/Brand or Model/Type cell (parsed by
    parse_inquiry) as the primary hint, then the description text. If nothing
    in inventory matches, the customer's raw text is kept so no info is lost."""
    for col in ("brand", "model"):
        if col not in df.columns:
            df[col] = ""
    if "item_no" in df.columns:      # parsers emit int or str; TextColumn needs str
        df["item_no"] = df["item_no"].astype(str)
    if df.empty:
        return df
    brands, models = inquiry_terms()
    for i, r in df.iterrows():
        b, m = extract_brand_model(r["description"], brands, models,
                                   brand_hint=r["brand"], model_hint=r["model"])
        df.at[i, "brand"] = b or r["brand"]
        df.at[i, "model"] = m or r["model"]
    return df


@st.cache_data
def best_match_preview(part_number, description, brand, model, cust_group):
    """Confidence + available qty of the single best inventory match for a
    customer line — shown on the customer grid so the team can eyeball
    quality/stock before drilling into any row. Matching is brand/model-scoped."""
    matches = match_row(get_con(), inquiry_corpus(), part_number, description,
                        brand=brand, model=model, cust_group=cust_group, limit=1)
    if matches.empty:
        return float("nan"), None   # blank cell, not a 0.0 that reads like a tier
    m = matches.iloc[0]
    return float(m["confidence"]), int(m["qty"])


def add_confidence_and_qty_match(grid):
    def _preview(r):
        cust_group = model_group_code(get_con(), r.get("model", ""))
        return best_match_preview(r["part_number"], r["description"],
                                  r.get("brand", ""), r.get("model", ""), cust_group)

    def _qty_match(avail, needed):
        if pd.isna(avail):
            return ""
        a = int(avail)
        if pd.isna(needed):        # freshly added row with no qty typed yet
            return f"{a}/?"
        n = int(needed)
        return f"{a}/{n} {'✅' if a >= n else '⚠️'}"

    preview = grid.apply(lambda r: pd.Series(_preview(r), index=["confidence", "_qty_avail"]), axis=1)
    grid["confidence"] = preview["confidence"]
    grid["qty_match"] = [_qty_match(avail, needed)
                         for avail, needed in zip(preview["_qty_avail"], grid["qty"])]
    return grid


def fuzzy_ids(text, limit=400):
    try:
        from rapidfuzz import process, fuzz
        ids, choices = fuzzy_corpus()
        hits = process.extract(text, choices, scorer=fuzz.WRatio,
                               limit=limit, score_cutoff=62)
        return [ids[i] for _, _, i in hits]
    except Exception:
        con = get_con()
        like = f"%{text}%"
        return [r[0] for r in con.execute(
            "SELECT row_id FROM parts WHERE part_name LIKE ? OR model LIKE ? LIMIT ?",
            (like, like, limit)).fetchall()]


def run_search(text, groups, brands, conds, cats, locs, min_stars, in_stock):
    con = get_con()
    where, params, fuzzy = [], [], False
    if text.strip():
        terms = " ".join(f'"{t}"*' for t in text.split())
        ids = [r[0] for r in con.execute(
            "SELECT rowid FROM parts_fts WHERE parts_fts MATCH ? LIMIT 4000",
            (terms,)).fetchall()]
        if not ids:                       # exact search empty -> fuzzy fallback
            ids, fuzzy = fuzzy_ids(text), True
        if not ids:
            return pd.DataFrame(), False
        where.append(f"row_id IN ({','.join('?'*len(ids))})")
        params += ids
    if groups:
        where.append(f"group_code IN ({','.join('?'*len(groups))})"); params += groups
    if brands:
        where.append(f"brand IN ({','.join('?'*len(brands))})"); params += brands
    if conds:
        where.append(f"condition IN ({','.join('?'*len(conds))})"); params += conds
    if cats:
        where.append(f"category IN ({','.join('?'*len(cats))})"); params += cats
    if locs:
        where.append(f"location IN ({','.join('?'*len(locs))})"); params += locs
    if min_stars > 1:
        where.append("stars >= ?"); params.append(min_stars)
    if in_stock:
        where.append("available = 1")
    sql = "SELECT * FROM parts"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY stars DESC, available DESC, part_name LIMIT 1500"
    return pd.read_sql_query(sql, con, params=params), fuzzy


def group_summary(group_code):
    con = get_con()
    return pd.read_sql_query(
        "SELECT * FROM parts WHERE group_code=? ORDER BY available DESC, stars DESC, part_name",
        con, params=(group_code,))


# ----------------------------------------------------------------- shortlist
def add_to_shortlist(r):
    sl = st.session_state.setdefault("shortlist", {})
    rid = int(r["row_id"])
    if rid not in sl:
        sl[rid] = {"part_name": r["part_name"], "brand": r["brand"], "model": r["model"],
                   "group_code": r["group_code"], "condition": r["condition"],
                   "qty_avail": int(r["qty"]), "unit": r["unit"], "location": r["location"],
                   "rack": r["rack"], "part_number": r["part_number"],
                   "stars": int(r["stars"]), "qty_needed": 1}


# ----------------------------------------------------------------- exports
EXPORT_COLS = {"part_name": "Part Name", "part_number": "Part No.", "brand": "Brand",
               "model": "Model", "group_code": "Group", "condition": "Condition",
               "qty": "Qty", "unit": "Unit", "location": "Location", "rack": "Rack",
               "details": "Details / Dimension", "stars": "Info ★", "available": "In Stock"}


def style_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="0F9D8F")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def to_excel(df, group_code):
    sub = df[list(EXPORT_COLS)].rename(columns=EXPORT_COLS).copy()
    sub["In Stock"] = sub["In Stock"].map({1: "Yes", 0: "No"})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        sub.to_excel(xl, index=False, sheet_name="Spares")
        ws = xl.sheets["Spares"]; style_header(ws)
        for i, w in enumerate([34, 16, 14, 14, 12, 14, 7, 8, 14, 12, 40, 8, 9], 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf.seek(0); return buf.getvalue()


def _pdf(rows, head, colw, title, subtitle):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12*mm,
                            rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    ss = getSampleStyleSheet()
    ts = ParagraphStyle("t", parent=ss["Title"], textColor=colors.HexColor(TEAL_DK), fontSize=18)
    sb = ParagraphStyle("s", parent=ss["Normal"], textColor=colors.HexColor(MUTE), fontSize=9)
    cell = ParagraphStyle("c", parent=ss["Normal"], fontSize=7, leading=8)
    flow = [Paragraph(title, ts)]
    for line in subtitle:
        flow.append(Paragraph(line, sb))
    flow.append(Spacer(1, 6*mm))
    data = [head] + [[Paragraph(str(c), cell) if not isinstance(c, int) else c for c in row]
                     for row in rows]
    t = Table(data, repeatRows=1, colWidths=[w*mm for w in colw])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(TEAL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1FAF8")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cfe4e0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 7)]))
    flow.append(t); doc.build(flow); buf.seek(0); return buf.getvalue()


def to_pdf(df, group_code, models):
    rows = [[r["part_name"], r["part_number"], r["brand"], r["model"], r["condition"],
             int(r["qty"]), r["location"], r["rack"], int(r["stars"])]
            for _, r in df.iterrows()]
    return _pdf(rows, ["Part Name", "Part No.", "Brand", "Model", "Condition",
                       "Qty", "Location", "Rack", "★"],
                [70, 26, 22, 24, 24, 12, 26, 30, 10],
                f"AbhiMarine — Spares for Model Group {group_code}",
                ["Interchangeable models: " + ", ".join(models[:25]) + (" …" if len(models) > 25 else ""),
                 f"{len(df)} line items · generated {datetime.now():%d %b %Y %H:%M}"])


def share_text(items):
    lines = ["*AbhiMarine — Parts enquiry*", ""]
    for v in items.values():
        pn = f" (PN {v['part_number']})" if v["part_number"] else ""
        lines.append(f"• {v['part_name']}{pn} — {v['brand']} {v['model']} · "
                     f"need {v['qty_needed']}, {v['qty_avail']} in stock · {v['condition']}")
    lines += ["", "AbhiMarine Pvt. Ltd. · abhimarine.com"]
    return "\n".join(lines)


# ----------------------------------------------------------------- header
con = get_con()
total = con.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
instock = con.execute("SELECT COUNT(*) FROM parts WHERE available=1").fetchone()[0]
ngroups = con.execute("SELECT COUNT(DISTINCT group_code) FROM parts WHERE group_code!=''").fetchone()[0]
nbrands = con.execute("SELECT COUNT(DISTINCT brand) FROM parts").fetchone()[0]

appbar = st.container(key="appbar")
top = appbar.columns([6, 1], vertical_alignment="center")
top[0].markdown('<div class="brand">⚓ <span>AbhiMarine</span> Parts Finder</div>',
                unsafe_allow_html=True)
if top[1].button("Sign out", use_container_width=True):
    st.session_state.clear(); st.rerun()

k = st.columns(4)
for col, v, l in zip(k, [f"{total:,}", f"{instock:,}", f"{ngroups:,}", f"{nbrands:,}"],
                     ["Total Parts", "In Stock", "Model Groups", "Brands"]):
    col.markdown(f"<div class='kpi'><div class='v'>{v}</div><div class='l'>{l}</div></div>",
                 unsafe_allow_html=True)

st.session_state.setdefault("group", None)
st.session_state.setdefault("shortlist", {})

# ----------------------------------------------------------------- top toolbar
opts = filter_options()
st.markdown('<div class="toolbar">', unsafe_allow_html=True)
tb1 = st.columns([3, 1.6, 1.4])
q = tb1[0].text_input("Search", placeholder="e.g.  L20 plunger  ·  MAN B&W scraper ring  ·  8525-784",
                      label_visibility="collapsed")
f_group = tb1[1].multiselect("Model Group", opts["groups"], placeholder="Model Group",
                             label_visibility="collapsed")
f_brand = tb1[2].multiselect("Brand", opts["brands"], placeholder="Brand", label_visibility="collapsed")
tb2 = st.columns([1.3, 1.3, 1.3, 1.6, 1.1])
f_cond = tb2[0].multiselect("Condition", opts["conditions"], placeholder="Condition", label_visibility="collapsed")
f_cat = tb2[1].multiselect("Category", opts["categories"], placeholder="Category", label_visibility="collapsed")
f_loc = tb2[2].multiselect("Location", opts["locations"], placeholder="Location", label_visibility="collapsed")
f_stars = tb2[3].slider("Min info ★", 1, 5, 1,
                        help="Stars rate how complete a listing is — part number, dimensions, "
                             "markings, genuinity and photo all add detail a buyer can trust.")
f_stock = tb2[4].checkbox("In stock", value=True)
st.markdown('</div>', unsafe_allow_html=True)

n_sl = len(st.session_state["shortlist"])
tab_search, tab_quote, tab_inquiry = st.tabs(
    ["🔎 Search", f"🧾 Enquiry / Quote ({n_sl})", "📥 Match Customer Inquiry"])

# ================================================================= SEARCH TAB
with tab_search:
    res, fuzzy = run_search(q, f_group, f_brand, f_cond, f_cat, f_loc, f_stars, f_stock)

    msg = f"**{len(res):,}** matching parts" + (" (showing first 1,500)" if len(res) == 1500 else "")
    if fuzzy:
        msg += "  ·  _no exact match — showing closest results_"
    st.markdown(msg)

    if res.empty:
        st.info("No parts match. Try fewer filters or a broader search term.")
    else:
        c_group, c_model = st.columns(2)

        # 1) pick a model group first — narrows both the model dropdown and the table below
        picked_group = None
        grp_counts = res[res["group_code"] != ""]["group_code"].value_counts().head(40)
        with c_group:
            if len(grp_counts):
                choices = ["— select a model group —"] + [f"{g}  ({n} here)" for g, n in grp_counts.items()]
                pick = st.selectbox("📦 Select a model group:", choices, key="grp_pick")
                picked_group = None if pick.startswith("—") else pick.split("  (")[0]
                st.session_state["group"] = picked_group
                if picked_group:
                    res = res[res["group_code"] == picked_group]

        # 2) model dropdown only lists models present within the (now group-filtered) results;
        # keyed per group so it resets cleanly when the group selection changes
        model_counts = res[res["model"] != ""]["model"].value_counts().head(60)
        with c_model:
            if len(model_counts):
                choices_m = ["— select a model —"] + [f"{m}  ({n} here)" for m, n in model_counts.items()]
                pick_m = st.selectbox("🔧 Select a model:", choices_m,
                                      key=f"model_pick_{picked_group or 'all'}")
                if not pick_m.startswith("—"):
                    res = res[res["model"] == pick_m.split("  (")[0]]

        grid = res.copy()
        grid.insert(0, "Add", False)
        grid["photo"] = grid["photo"].map({1: "📷", 0: ""})
        grid["available"] = grid["available"].map({1: "✅ In stock", 0: "❌ Out"})

        edited = st.data_editor(
            grid,
            column_order=["Add", "sr", "photo", "part_name", "brand", "model", "group_code",
                         "condition", "qty", "unit", "location", "rack", "part_number",
                         "stars", "available", "details"],
            column_config={
                "Add": st.column_config.CheckboxColumn("➕", width="small"),
                "sr": st.column_config.TextColumn("SR No.", width="small"),
                "photo": st.column_config.TextColumn("📷", width="small"),
                "part_name": st.column_config.TextColumn("Part Name", width="large"),
                "brand": st.column_config.TextColumn("Brand"),
                "model": st.column_config.TextColumn("Model"),
                "group_code": st.column_config.TextColumn("Group"),
                "condition": st.column_config.TextColumn("Condition"),
                "qty": st.column_config.NumberColumn("Qty", format="%d"),
                "unit": st.column_config.TextColumn("Unit", width="small"),
                "location": st.column_config.TextColumn("Location"),
                "rack": st.column_config.TextColumn("Rack"),
                "part_number": st.column_config.TextColumn("Part No."),
                "stars": st.column_config.NumberColumn("★", format="%d ★"),
                "available": st.column_config.TextColumn("Availability"),
                "details": st.column_config.TextColumn("Details / Dimension", width="large"),
            },
            disabled=[c for c in grid.columns if c != "Add"],
            hide_index=True,
            use_container_width=True,
            height=560,
            key="results_grid",
        )

        checked = edited[edited["Add"]]
        if len(checked) and st.button(f"➕ Add {len(checked)} checked row(s) to enquiry"):
            for rid in checked["row_id"]:
                add_to_shortlist(res[res["row_id"] == rid].iloc[0])
            st.rerun()

# ================================================================= QUOTE TAB
with tab_quote:
    sl = st.session_state["shortlist"]
    if not sl:
        st.info("Your enquiry list is empty. Add parts from the Search tab with the ➕ button.")
    else:
        st.markdown(f"### 🧾 Enquiry list — {len(sl)} part(s)")
        for rid, v in list(sl.items()):
            c1, c2, c3 = st.columns([8, 2, 1])
            c1.markdown(f"**{v['part_name']}** · {v['brand']} {v['model']} · "
                        f"{v['condition']} · {v['qty_avail']} in stock"
                        f"{' · PN ' + v['part_number'] if v['part_number'] else ''}")
            v["qty_needed"] = c2.number_input("Qty needed", min_value=1, value=v["qty_needed"],
                                              key=f"qn_{rid}", label_visibility="collapsed")
            if c3.button("🗑", key=f"rm_{rid}", help="Remove"):
                del sl[rid]; st.rerun()

        st.divider()
        e1, e2, e3, e4 = st.columns([1, 1, 1, 1])
        e1.download_button("⬇ Excel quote", build_estimate_xlsx(sl),
                           file_name=f"Estimate_{datetime.now():%Y%m%d}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
        try:
            e2.download_button("⬇ PDF quote", build_estimate_pdf(sl),
                               file_name=f"Estimate_{datetime.now():%Y%m%d}.pdf",
                               mime="application/pdf", use_container_width=True)
        except Exception:
            e2.caption("PDF: `pip install reportlab`")
        txt = share_text(sl)
        e3.link_button("💬 WhatsApp", "https://wa.me/?text=" + urllib.parse.quote(txt),
                       use_container_width=True)
        e4.link_button("✉ Email", "mailto:?subject=" + urllib.parse.quote("AbhiMarine parts enquiry") +
                       "&body=" + urllib.parse.quote(txt), use_container_width=True)
        if st.button("Clear list"):
            st.session_state["shortlist"] = {}; st.rerun()
        with st.expander("Preview share message"):
            st.code(txt)

# ================================================================= INQUIRY TAB
with tab_inquiry:
    st.markdown("Dump a customer inquiry here — **Excel, Word, PDF, CSV, or a photo/scan** — "
               "and it'll be turned into a part list. Tick **➕** next to any line to see "
               "matching inventory on the right, then add matches straight to the enquiry.")
    up = st.file_uploader("Customer inquiry file", label_visibility="collapsed",
                          type=["xlsx", "xls", "xlsm", "csv", "doc", "docx", "pdf",
                                "jpg", "jpeg", "png"])

    if up is not None:
        file_key = f"{up.name}_{up.size}"
        if st.session_state.get("inquiry_file_key") != file_key:
            with st.spinner("Reading file…"):
                try:
                    parsed = parse_inquiry(up.name, up.getvalue())
                    if parsed.empty:
                        st.warning("Couldn't find a recognisable part list in this file — "
                                  "you can still add rows manually below.")
                except Exception as e:
                    st.error(f"Couldn't read this file: {e}")
                    parsed = pd.DataFrame(columns=["item_no", "description", "part_number", "qty"])
            st.session_state["inquiry_df"] = annotate_brand_model(parsed)
            st.session_state["inquiry_file_key"] = file_key

    with st.expander("Or paste text — copied from Excel, a website, or a PDF"):
        pasted = st.text_area("Paste customer inquiry text", height=150,
                              label_visibility="collapsed", key="inquiry_paste_box",
                              placeholder="Paste rows here, e.g.:\nNo\tDescription\tPart No\tQty\n"
                                         "1\tF.O. Duplex Filter Bush\t51403-03D-20\t12")
        if st.button("Parse pasted text") and pasted.strip():
            parsed = parse_pasted_text(pasted)
            if parsed.empty:
                st.warning("Couldn't find a recognisable part list in the pasted text — "
                          "you can still add rows manually below.")
            st.session_state["inquiry_df"] = annotate_brand_model(parsed)
            st.session_state["inquiry_file_key"] = f"paste_{hash(pasted)}"

    idf = st.session_state.get("inquiry_df")
    if idf is None:
        st.info("Upload a file above to extract the customer's part list.")
    else:
        # Keep the Match ticks inside the stored frame so they survive the rerun
        # we trigger on every edit. Confidence + qty-match are recomputed each run
        # so they refresh live once the edited values are persisted below.
        if "Match" not in idf.columns:
            idf.insert(0, "Match", False)
            st.session_state["inquiry_df"] = idf
        grid = add_confidence_and_qty_match(idf.copy())

        # caption and "Matching inventory" share one header row, so the title
        # sits on the same line as the left caption and both tables start level
        h_l, h_r = st.columns([1.4, 1])
        h_l.caption(f"Extracted **{len(idf)}** line item(s) — fix any misread cell below "
                    "(common with photos/scans) before matching.")
        h_r.markdown(f"<div style='font-weight:700;color:{INK};font-size:1.05rem'>"
                     "Matching inventory</div>", unsafe_allow_html=True)

        col_l, col_r = st.columns([1.4, 1])
        wkey = (f"inquiry_grid_{st.session_state.get('inquiry_file_key', '')}"
                f"_{st.session_state.get('grid_ver', 0)}")
        with col_l:
            i_edited = st.data_editor(
                grid,
                column_order=["Match", "item_no", "description", "part_number", "qty",
                             "brand", "model", "confidence", "qty_match"],
                column_config={
                    "Match": st.column_config.CheckboxColumn("➕", width="small"),
                    "item_no": st.column_config.TextColumn("No.", width="small"),
                    "description": st.column_config.TextColumn("Description", width="large"),
                    "part_number": st.column_config.TextColumn("Part No."),
                    "qty": st.column_config.NumberColumn("Qty"),
                    "brand": st.column_config.TextColumn("Brand"),
                    "model": st.column_config.TextColumn("Model"),
                    "confidence": st.column_config.NumberColumn("Confidence (CL)", format="%.1f"),
                    "qty_match": st.column_config.TextColumn("Qty Match", width="small"),
                },
                disabled=["confidence", "qty_match"],
                num_rows="dynamic",
                hide_index=True,
                use_container_width=True,
                height=520,
                key=wkey,
            )

        # Persist edits/adds/deletes so confidence, qty-match and the right-hand
        # matches all recompute from the edited values on the next run. Only a
        # structural change (row added/deleted) bumps grid_ver to reset the
        # dynamic-row editor — otherwise its delta would re-insert added rows.
        keep = [c for c in ["Match", "item_no", "description", "part_number",
                            "qty", "brand", "model"] if c in i_edited.columns]
        fed = grid[keep].reset_index(drop=True)
        got = i_edited[keep].reset_index(drop=True)
        try:
            got = got.astype(fed.dtypes.to_dict())   # ignore pure dtype drift
        except Exception:
            pass
        if not got.equals(fed):
            st.session_state["inquiry_df"] = got
            delta = st.session_state.get(wkey, {})
            if delta.get("added_rows") or delta.get("deleted_rows"):
                st.session_state["grid_ver"] = st.session_state.get("grid_ver", 0) + 1
            st.rerun()

        with col_r, st.container(key="matchcol"):
            checked = i_edited[i_edited["Match"] == True]
            if checked.empty:
                st.caption("Tick ➕ next to a line on the left to see matches here.")
            else:
                corpus = inquiry_corpus()
                for idx, row in checked.iterrows():
                    label = row["description"] or row["part_number"] or f"Item {row['item_no']}"
                    cust_group = model_group_code(get_con(), row.get("model", ""))
                    with st.expander(f"🔎 {label}  ·  need {row['qty']}", expanded=True):
                        matches = match_row(get_con(), corpus, row["part_number"], row["description"],
                                            brand=row.get("brand", ""), model=row.get("model", ""),
                                            cust_group=cust_group)
                        if matches.empty:
                            st.warning("No matching part found in inventory.")
                        else:
                            mgrid = matches.copy()
                            mgrid["available"] = mgrid["available"].astype(bool)
                            mgrid["photo"] = mgrid["photo"].astype(bool)
                            mgrid.insert(0, "Add", False)
                            m_edited = st.data_editor(
                                mgrid,
                                column_order=["Add", "confidence", "match_type", "part_name", "brand",
                                             "model", "condition", "qty", "location", "rack",
                                             "sr", "part_number", "photo", "available"],
                                column_config={
                                    "Add": st.column_config.CheckboxColumn("➕", width="small"),
                                    "confidence": st.column_config.NumberColumn("Confidence (CL)", format="%.1f"),
                                    "match_type": st.column_config.TextColumn("Match"),
                                    "part_name": st.column_config.TextColumn("Part Name", width="large"),
                                    "qty": st.column_config.NumberColumn("Qty avail", format="%d"),
                                    "sr": st.column_config.TextColumn("Serial No."),
                                    "photo": st.column_config.CheckboxColumn("Photo"),
                                    "available": st.column_config.CheckboxColumn("In stock"),
                                },
                                disabled=[c for c in mgrid.columns if c != "Add"],
                                hide_index=True,
                                use_container_width=True,
                                key=f"match_grid_{idx}",
                            )
                            m_checked = m_edited[m_edited["Add"] == True]
                            if len(m_checked) and st.button(
                                    f"➕ Add {len(m_checked)} to enquiry", key=f"addm_{idx}"):
                                for rid in m_checked["row_id"]:
                                    add_to_shortlist(matches[matches["row_id"] == rid].iloc[0])
                                st.rerun()

# ----------------------------------------------------------------- bottom bar
g = st.session_state.get("group")
if g:
    gdf = group_summary(g)
    avail_df = gdf[gdf["available"] == 1]
    models = models_in_group(g)
    tot_qty = int(avail_df["qty"].sum())

    st.markdown(f"""<div class="footbar">
    Model group <b>{g}</b> — <b>{len(avail_df)}</b> spares in stock
    · total qty <b>{tot_qty}</b> · interchangeable across <b>{len(models)}</b> models</div>
    """, unsafe_allow_html=True)

    with st.expander(f"📦  {g} — full spares list & export", expanded=False):
        st.markdown("**Interchangeable models:** " +
                    " ".join(f"<span class='pill'>{m}</span>" for m in models),
                    unsafe_allow_html=True)
        c1, c2, c3 = st.columns([1, 1, 4])
        c1.download_button("⬇ Excel", to_excel(gdf, g),
                           file_name=f"AbhiMarine_{g.replace('/','-')}_spares.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
        try:
            c2.download_button("⬇ PDF", to_pdf(gdf, g, models),
                               file_name=f"AbhiMarine_{g.replace('/','-')}_spares.pdf",
                               mime="application/pdf", use_container_width=True)
        except Exception:
            c2.caption("PDF: `pip install reportlab`")
        c3.caption(f"{len(gdf)} total line items · {len(avail_df)} in stock")
        st.dataframe(
            gdf[["sr", "part_name", "brand", "model", "condition", "qty", "unit",
                 "location", "rack", "part_number", "stars", "available"]]
            .rename(columns={"sr": "SR No.", "available": "in_stock"}),
            use_container_width=True, hide_index=True)
