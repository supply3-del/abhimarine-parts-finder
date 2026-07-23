"""
Quotation export in the Excel International "Estimate" house style
(matches the sample files in ../sample quotes/). Line items are filled from
the enquiry shortlist; price and customer cells are left blank for the team to
complete by hand, exactly like the source template.
"""
import io
import os
from datetime import datetime

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "excel_intl_logo.png")
WAREHOUSE = "GATE-1-WAREHOUSE NO: C1-25, AJMAN FREE ZONE, AJMAN 5908, UAE"

# standard Excel International terms (the two job-specific ones are left as blanks to fill)
TERMS = [
    "Delivery time :   ____ weeks after receipt of payment",
    "Payment terms: ____________________________________",
    "Dismantling if any additional, Packing, Delivery & VAT charges will be extra",
    "For orders below USD 500, small order surcharge of USD 100 will be applicable.",
    "We may need more details at the time of order confirmation.",
    "5% VAT will be charged if delivery will not be directly to the vessel",
    "Delivery Times are indicative from ex works and are not a condition of any contract or order.",
]
HEADERS = ["Sr", "Part name", "Condition", "Ex works", "UOM", "Qty", "Rate (USD)", "Amount (USD)"]

INK = "1B2A32"
BLU = "2E6E7E"      # excel-international teal-blue
LINE = "9BB4BC"


def _uniform(items, key):
    """Return the shared value of `key` across all lines, else '' (so Make/Model
    only auto-fill when the whole quote is one brand/model)."""
    vals = {(v.get(key) or "").strip() for v in items.values()}
    vals.discard("")
    return next(iter(vals)) if len(vals) == 1 else ""


def _part_label(v):
    pn = (v.get("part_number") or "").strip()
    return v["part_name"] + (f"   (Part No: {pn})" if pn else "")


# ----------------------------------------------------------------- Excel
def build_estimate_xlsx(items) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"
    ws.sheet_view.showGridLines = False

    widths = {"A": 2.5, "B": 8, "C": 52, "D": 16, "E": 12, "F": 8, "G": 8, "H": 14, "I": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    thin = Side(style="thin", color=LINE)
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width, img.height = 300, 74
        ws.add_image(img, "B1")

    def put(coord, val, *, bold=False, size=11, color=INK, align=None, fill=None):
        c = ws[coord]
        c.value = val
        c.font = Font(bold=bold, size=size, color=color, name="Calibri")
        if align:
            c.alignment = align
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        return c

    # title + meta block (right)
    ws.merge_cells("G6:I6"); put("G6", "Estimate", bold=True, size=24, color=BLU, align=center)
    put("G7", "Ref no", bold=True); ws.merge_cells("H7:I7"); put("H7", "", align=left)
    put("G8", "Date", bold=True); ws.merge_cells("H8:I8")
    put("H8", f"{datetime.now():%d-%m-%Y}", align=left)
    put("G9", "Make", bold=True); ws.merge_cells("H9:I9"); put("H9", _uniform(items, "brand"), align=left)
    put("G10", "Model", bold=True); ws.merge_cells("H10:I10"); put("H10", _uniform(items, "model"), align=left)

    # address + bill-to (left)
    ws.merge_cells("B9:E9"); put("B9", WAREHOUSE, size=10, align=left)
    put("B11", " Bill To.", bold=True, size=12)
    ws.merge_cells("B12:E12"); put("B12", "", align=left)
    put("B13", " Kind Attention", bold=True)
    ws.merge_cells("B14:E14"); put("B14", "", align=left)

    # table header
    hdr = 15
    for j, h in enumerate(HEADERS):
        c = put(ws.cell(hdr, 2 + j).coordinate, h, bold=True, size=11, color="FFFFFF",
                align=center, fill=BLU)
        c.border = box

    # rows
    r = hdr + 1
    for i, v in enumerate(items.values(), 1):
        vals = [i, _part_label(v), v.get("condition", ""), "", v.get("unit", ""),
                v.get("qty_needed", 1), "", None]
        for j, val in enumerate(vals):
            c = ws.cell(r, 2 + j)
            c.value = val
            c.border = box
            c.alignment = left if j == 1 else center
            c.font = Font(size=11, color=INK, name="Calibri")
        ws.cell(r, 9).value = f'=IF(H{r}="","",H{r}*G{r})'   # Amount = Rate*Qty, blank until priced
        r += 1

    # total
    ws.merge_cells(f"B{r}:H{r}")
    put(f"B{r}", "TOTAL USD", bold=True, size=12, align=Alignment(horizontal="right", vertical="center"),
        fill="EAF1F3")
    tot = put(f"I{r}", f'=IF(SUM(I{hdr+1}:I{r-1})=0,"",SUM(I{hdr+1}:I{r-1}))',
              bold=True, size=12, align=center, fill="EAF1F3")
    for col in "BCDEFGHI":
        ws[f"{col}{r}"].border = box

    # terms
    r += 2
    put(f"B{r}", "Terms & Conditions:", bold=True, size=11)
    for term in TERMS:
        r += 1
        put(f"B{r}", term, size=10, align=left)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ----------------------------------------------------------------- PDF
def build_estimate_pdf(items) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, Image)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    ss = getSampleStyleSheet()
    blu = colors.HexColor("#" + BLU)
    small = ParagraphStyle("s", parent=ss["Normal"], fontSize=8, leading=10)
    cell = ParagraphStyle("c", parent=ss["Normal"], fontSize=8, leading=9)
    lab = ParagraphStyle("l", parent=ss["Normal"], fontSize=9, leading=12, textColor=blu, fontName="Helvetica-Bold")
    flow = []

    # header: logo (left) + Estimate/meta (right)
    meta = [["Estimate", ""], ["Ref no", ""], ["Date", f"{datetime.now():%d-%m-%Y}"],
            ["Make", _uniform(items, "brand")], ["Model", _uniform(items, "model")]]
    meta_tbl = Table(meta, colWidths=[24 * mm, 44 * mm])
    meta_tbl.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)), ("FONTSIZE", (0, 0), (1, 0), 20),
        ("TEXTCOLOR", (0, 0), (1, 0), blu), ("ALIGN", (0, 0), (1, 0), "RIGHT"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"), ("TEXTCOLOR", (0, 1), (0, -1), blu),
        ("FONTSIZE", (0, 1), (-1, -1), 9), ("ALIGN", (0, 1), (0, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2), ("TOPPADDING", (0, 0), (-1, -1), 2)]))
    logo = Image(LOGO_PATH, width=62 * mm, height=15 * mm) if os.path.exists(LOGO_PATH) else Paragraph("", small)
    head = Table([[logo, meta_tbl]], colWidths=[95 * mm, 87 * mm])
    head.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    flow += [head, Spacer(1, 4 * mm)]

    flow += [Paragraph(WAREHOUSE, small), Spacer(1, 3 * mm),
             Paragraph("<b>Bill To.</b>", small), Paragraph("&nbsp;", small),
             Paragraph("<b>Kind Attention</b>", small), Paragraph("&nbsp;", small),
             Spacer(1, 3 * mm)]

    # items table
    data = [HEADERS]
    for i, v in enumerate(items.values(), 1):
        data.append([str(i), Paragraph(_part_label(v), cell), v.get("condition", ""), "",
                     v.get("unit", ""), str(v.get("qty_needed", 1)), "", ""])
    data.append(["", "TOTAL USD", "", "", "", "", "", ""])
    colw = [9, 57, 19, 15, 11, 11, 21, 23]   # sums to 166mm < printable width (182mm)
    t = Table(data, repeatRows=1, colWidths=[w * mm for w in colw])
    n = len(data) - 1
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), blu), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, n - 1), 0.4, colors.HexColor("#" + LINE)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ("SPAN", (1, n), (6, n)), ("ALIGN", (1, n), (6, n), "RIGHT"),
        ("FONTNAME", (1, n), (-1, n), "Helvetica-Bold"), ("BACKGROUND", (0, n), (-1, n), colors.HexColor("#EAF1F3")),
        ("BOX", (0, n), (-1, n), 0.4, colors.HexColor("#" + LINE))]))
    flow += [t, Spacer(1, 5 * mm)]

    flow.append(Paragraph("Terms &amp; Conditions:", lab))
    for term in TERMS:
        flow.append(Paragraph(term.replace("&", "&amp;"), small))

    doc.build(flow); buf.seek(0)
    return buf.getvalue()
